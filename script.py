import pandas as pd
import openpyxl
import sys

wb = openpyxl.load_workbook('export.xlsx')
ws = wb.active
max_col = ws.max_column
max_row = ws.max_row
print(max_row)


def prepare_xlsx():
    ws.delete_cols(1, 1)
    ws.delete_cols(5, 3)
    wb.save('mode.xlsx')


def prepare_precond(s):
    if s > 0:
        ws.move_range(f'C2:C{s + 1}', cols=-1)
    ws.cell(1, 2).value = "Предусловие"
    for i in range(2, s + 1):
        ws[f'D{i}'].value = None
    wb.save('mode.xlsx')


def prepare_precondition(s):
    if s > 0:
        s += 1
        ws.cell(1, 2).value = "Предусловие"
        print(f'"C2:C{s}"')
        for i in range(2, s + 1):
            cell_obj = ws.cell(row=i, column=3)
            ws[f'B{i}'].value = cell_obj.value
            ws[f'C{i}'].value = None
            ws[f'D{i}'].value = None
            ##ws.move_range(f'"C2:C3"', cols=-1)
    else:
        ws.delete_cols(2, 1)
    wb.save('mode.xlsx')


def concat_precondition(row_value):
    values = []
    del values[:]
    for row in ws.iter_cols(min_col=2, max_col=2, min_row=2, max_row=row_value):
        for cell in row:
            if cell.value is None:
                pass
            else:
                values.append(str(cell.value))

    ws[f'B2'].value = '\n'.join(values)
    for row in range(2, row_value):
        ws[f'B{row + 1}'].value = None
    wb.save('mode.xlsx')


def concat_step(row_value):
    values = []
    del values[:]
    for row in ws.iter_cols(min_col=3, max_col=3, min_row=2, max_row=row_value):
        for cell in row:
            if cell.value is None:
                pass
            else:
                values.append(str(cell.value))

    ws[f'C2'].value = ';'.join(values)
    for row in range(2, row_value):
        ws[f'C{row + 1}'].value = None
    wb.save('mode.xlsx')


def concat_step_with_actual_result(row_value, s):
    values = []
    del values[:]
    values_ = ''
    for col in ws.iter_rows(min_col=3, max_col=4, min_row=2, max_row=row_value):  # values_only=False
        for cell in col:
            if cell.value is None:
                pass
            else:
                if values_ == '':
                    values_ += str(cell.value)
                else:
                    ##values_ += ('\nОР:\n' + str(cell.value))
                    values_ += f'\nОР: \n {cell.value}'
        if values_ != '':
            values.append(values_)
        values_ = ''

    ws[f'C2'].value = ';'.join(values)
    for row in range(2, row_value):
        ws[f'C{row + 1}'].value = None
    if s == 0:
        ws.delete_cols(3, 1)
    else:
        ws.delete_cols(4, 1)
    wb.save('mode.xlsx')


def concat_tags(row_value):
    values = []
    del values[:]
    for row in ws.iter_cols(min_col=4, max_col=4, min_row=2, max_row=row_value):
        for cell in row:
            if cell.value is None:
                pass
            else:
                values.append(str(cell.value))

    ws[f'D2'].value = ';'.join(values)
    for row in range(2, row_value):
        ws[f'D{row + 1}'].value = None
    wb.save('mode.xlsx')


def export_to_csv():
    df = pd.read_excel(r'mode.xlsx')
    df.to_csv('export.csv', index=False, header=True)


if __name__ == '__main__':
    prepare_xlsx()
    ##prepare_precondition()
    prepare_precond(int(sys.argv[1]))
    ##prepare_precondition(int(sys.argv[1]))
    concat_precondition(max_row)
    ##concat_step(max_row)
    concat_step_with_actual_result(max_row, int(sys.argv[1]))
    concat_tags(max_row)
    export_to_csv()
